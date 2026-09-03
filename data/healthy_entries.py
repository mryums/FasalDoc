# -*- coding: utf-8 -*-
"""Build all outputs: normalized image folder, enhanced Excel, JSON, image mapping CSV.

Usage:  python build_dataset.py
Reads:  dataset/*.py, C:/Users/DELL/Downloads/demo photos/*
Writes: plant_disease_dataset/{crop_knowledge_enhanced.xlsx, crop_knowledge.json,
        image_mapping.csv, images/...}
"""
import csv
import importlib.util
import json
import re
import shutil
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent
SRC_PHOTOS = Path(r"C:\Users\DELL\Downloads\demo photos")
OUT = BASE / "plant_disease_dataset"
IMG_OUT = OUT / "images"

# ---------------------------------------------------------------- load data
def load(name, path):
    spec = importlib.util.spec_from_file_location(name, path)
    m = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(m)
    return m

p1 = load("p1", BASE / "dataset" / "part1_vegetables_cereals.py")
p2 = load("p2", BASE / "dataset" / "part2_rice_cotton_chili_onion.py")
p3 = load("p3", BASE / "dataset" / "part3_fruits_legume.py")
pl = load("pl", BASE / "dataset" / "plants.py")
he = load("he", BASE / "dataset" / "healthy_entries.py")

DISEASES = p1.DISEASES + p2.DISEASES + p3.DISEASES
PLANTS = pl.PLANTS
PLANT_SLUGS = pl.PLANT_SLUGS
FOLDER_TO_PLANT = pl.FOLDER_TO_PLANT
HEALTHY = {p: (desc, tip) for p, desc, tip in he.HEALTHY_ENTRIES}  # plant -> (ur_description, ur_care_tip)

# (plant, disease) -> url-friendly slug; values MUST match FILENAME_TO_SLUG below
SLUG_BY_DISEASE = {
    ("Tomato", "Early Blight"): "early_blight",
    ("Tomato", "Late Blight"): "late_blight",
    ("Tomato", "Yellow Leaf Curl Virus"): "yellow_leaf_curl_virus",
    ("Tomato", "Septoria Leaf Spot"): "septoria_leaf_spot",
    ("Tomato", "Bacterial Spot"): "bacterial_spot",
    ("Potato", "Early Blight"): "early_blight",
    ("Potato", "Late Blight"): "late_blight",
    ("Potato", "Blackleg"): "blackleg",
    ("Potato", "Potato Virus Y (PVY)"): "potato_virus_y",
    ("Maize", "Northern Corn Leaf Blight"): "northern_corn_leaf_blight",
    ("Maize", "Common Rust"): "common_rust",
    ("Maize", "Gray Leaf Spot"): "gray_leaf_spot",
    ("Wheat", "Leaf Rust (Brown Rust)"): "leaf_rust",
    ("Wheat", "Yellow Rust (Stripe Rust)"): "yellow_rust",
    ("Wheat", "Powdery Mildew"): "powdery_mildew",
    ("Wheat", "Septoria Leaf Blotch"): "septoria_leaf_blotch",
    ("Rice", "Bacterial Leaf Blight"): "bacterial_leaf_blight",
    ("Rice", "Brown Spot"): "brown_spot",
    ("Rice", "Leaf Smut"): "leaf_smut",
    ("Rice", "Rice Blast"): "rice_blast",
    ("Rice", "Sheath Blight"): "sheath_blight",
    ("Cotton", "Alternaria Leaf Spot"): "alternaria_leaf_spot",
    ("Cotton", "Bacterial Blight (Angular Leaf Spot)"): "bacterial_blight",
    ("Cotton", "Stemphylium Leaf Spot"): "stemphylium_leaf_spot",
    ("Cotton", "Anthracnose of Cotton"): "anthracnose",
    ("Cotton", "Ascochyta Blight (Wet Weather Blight)"): "ascochyta_blight",
    ("Chili", "Leaf Curl Virus"): "leaf_curl_virus",
    ("Chili", "Anthracnose (Fruit Rot)"): "anthracnose_fruit_rot",
    ("Chili", "Powdery Mildew"): "powdery_mildew",
    ("Onion", "Purple Blotch"): "purple_blotch",
    ("Onion", "Downy Mildew"): "downy_mildew",
    ("Onion", "Basal Rot"): "basal_rot",
    ("Onion", "Powdery Mildew"): "powdery_mildew",
    ("Mango", "Anthracnose"): "anthracnose",
    ("Mango", "Powdery Mildew"): "powdery_mildew",
    ("Mango", "Bacterial Canker"): "bacterial_canker",
    ("Mango", "Sooty Mold"): "sooty_mold",
    ("Citrus", "Citrus Canker"): "citrus_canker",
    ("Citrus", "Citrus Greening (HLB)"): "citrus_greening",
    ("Citrus", "Melanose"): "melanose",
    ("Citrus", "Citrus Scab"): "citrus_scab",
    ("Citrus", "Black Spot"): "black_spot",
    ("Guava", "Algal Leaf Spot"): "algal_leaf_spot",
    ("Guava", "Anthracnose"): "anthracnose",
    ("Guava", "Fruit Scab"): "fruit_scab",
    ("Guava", "Stylar End Rot"): "stylar_end_rot",
    ("Guava", "Rust (Puccinia psidii)"): "rust",
    ("Chickpea", "Ascochyta Blight"): "ascochyta_blight",
    ("Chickpea", "Fusarium Wilt"): "fusarium_wilt",
    ("Chickpea", "Botrytis Grey Mold"): "botrytis_grey_mold",
}

for d in DISEASES:
    key = (d["plant"], d["disease"])
    if key not in SLUG_BY_DISEASE:
        raise SystemExit(f"Missing slug mapping for {key}")
    d["slug_disease"] = SLUG_BY_DISEASE[key]

# ---------------------------------------------------------------- assign IDs
plant_order = []  # canonical plant order from original sheet
for d in DISEASES:
    if d["plant"] not in plant_order:
        plant_order.append(d["plant"])

plant_ids = {p: f"PL{i+1:03d}" for i, p in enumerate(plant_order)}
ds_counter = 0
for d in DISEASES:
    ds_counter += 1
    d["disease_id"] = f"DS{ds_counter:03d}"

# healthy entries after all diseases, in plant order
HEALTHY_IDS = {}
for p in plant_order:
    ds_counter += 1
    HEALTHY_IDS[p] = f"DS{ds_counter:03d}"

# ---------------------------------------------------------------- image mapping
FILENAME_TO_SLUG = {
    "Tomato": {
        "Bacterial_Spot": "bacterial_spot", "Early_Blight": "early_blight",
        "Late_Blight": "late_blight", "Septoria_Leaf_Spot": "septoria_leaf_spot",
        "Yellow_Leaf_Curl_Virus": "yellow_leaf_curl_virus", "Healthy": "healthy",
    },
    "Potato": {
        "Blackleg": "blackleg", "Early_Blight": "early_blight", "Late_Blight": "late_blight",
        "Virus_Y": "potato_virus_y", "Healthy": "healthy",
    },
    "Maize": {
        "Common_Rust": "common_rust", "Gray_leaf_spot": "gray_leaf_spot",
        "Northern_Corn_Leaf_Blight": "northern_corn_leaf_blight", "Healthy": "healthy",
    },
    "Wheat": {
        "Brown_Rust": "leaf_rust", "Mildew": "powdery_mildew", "Septoria": "septoria_leaf_blotch",
        "Yellow_Rust": "yellow_rust", "Healthy": "healthy",
    },
    "Rice": {
        "Bacterial_Leaf_Blight": "bacterial_leaf_blight", "Blast": "rice_blast",
        "Brown_Spot": "brown_spot", "Sheath_Blight": "sheath_blight",
        "leaf_Smut": "leaf_smut", "Healthy": "healthy",
    },
    "Cotton": {
        "Alternaria_Leaf_Spot": "alternaria_leaf_spot", "Anthracnose": "anthracnose",
        "Ascochyta Blight": "ascochyta_blight", "Bacterial_Blight": "bacterial_blight",
        "Stemphylium_Leaf_Spot": "stemphylium_leaf_spot", "Healthy": "healthy",
    },
    "Chilli": {
        "Anthracnose (Fruit Rot)": "anthracnose_fruit_rot", "Leaf_Curl_Virus": "leaf_curl_virus",
        "Powdery_Mildew": "powdery_mildew", "Healthy": "healthy",
    },
    "Onion": {
        "Basal_Rot": "basal_rot", "Powdery_Mildew": "powdery_mildew",
        "Purple_Blotch": "purple_blotch", "Healthy": "healthy",
    },
    "Mango": {
        "Anthracnose": "anthracnose", "Bacterial_Canker": "bacterial_canker",
        "Powdery_Mildew": "powdery_mildew", "Sooty_Mold": "sooty_mold", "Healthy": "healthy",
    },
    "Citrus": {
        "Black_Spot": "black_spot", "Canker": "citrus_canker", "Greening": "citrus_greening",
        "Melanose": "melanose", "Scab": "citrus_scab", "Healthy": "healthy",
    },
    "Guava": {
        "Algal_Leaf_Spot": "algal_leaf_spot", "Anthracnose": "anthracnose",
        "Fruit_Scab": "fruit_scab", "Fungal_Rust": "rust",
        "Styler_End_Root": "stylar_end_rot", "Healthy": "healthy",
    },
    "Chickpea": {
        "Ascochyta_Blight": "ascochyta_blight", "Botrytis_Grey_Mold": "botrytis_grey_mold",
        "Fusarium_Wilt": "fusarium_wilt", "Healthy": "healthy",
    },
}

REVIEW_NOTES = {
    ("Chilli", "Anthracnose (Fruit Rot)", "01"): "Sampled photo shows water-damaged leaves rather than fruit-rot symptoms, and carries a Shutterstock watermark. Replace with a royalty-free photo of anthracnose fruit rot.",
    ("Onion", "Basal_Rot", "01"): "Sampled photo shows garlic, not onion. Replace with a true onion basal-rot photo or relabel.",
    ("Guava", "Fruit_Scab", "01"): "Sampled photo appears to be an anthracnose diagnostic collage (sunken lesions + Colletotrichum spores), not corky scab. Verify or replace.",
    ("Wheat", "Mildew", "01"): "Sampled photo shows a wetter/standing-water setting atypical for wheat. Verify the crop in the photo.",
    ("Citrus", "Greening", "01"): "Sampled photo shows normal green fruit with no clear HLB symptoms (blotchy leaf mottle). Verify or replace.",
}

def norm_ext(ext):
    ext = ext.lower()
    return "jpg" if ext in ("jpg", "jpeg", "jfif") else ext

# plant -> slug -> list of (num, new_path, label_src)
image_records = []   # dicts: id, path, original, folder, plant, disease_slug, num, status, review
unmapped_files = []

if SRC_PHOTOS.exists():
    raw = []
    for folder in sorted(SRC_PHOTOS.iterdir()):
        if not folder.is_dir():
            continue
        plant = FOLDER_TO_PLANT.get(folder.name)
        if plant is None:
            unmapped_files.append((folder.name, "unknown folder"))
            continue
        slug_map = FILENAME_TO_SLUG.get(folder.name, {})
        for f in sorted(folder.iterdir()):
            if not f.is_file():
                continue
            m = re.match(r"^(.*)_(\d{2})\.([A-Za-z]+)$", f.name)
            if m:
                base, num, ext = m.group(1), m.group(2), norm_ext(m.group(3))
            else:
                dot = f.name.rfind(".")
                base, num, ext = f.name[:dot], None, norm_ext(f.name[dot+1:])
            # base is "<CropPrefix>_<DiseaseKey>" or "Healthy_<Crop>"
            slug = None
            if base == f"Healthy_{folder.name}":
                slug = "healthy"
            else:
                for key, s in slug_map.items():
                    if s == "healthy":
                        continue
                    if base == f"{folder.name}_{key}":
                        slug = s
                        break
            if slug is None:
                unmapped_files.append((str(f), "filename not recognised"))
                continue
            raw.append({
                "folder": folder.name, "plant": plant, "slug": slug,
                "num": num or "01", "ext": ext, "original": str(f),
                "orig_name": f.name,
            })

    # order: plant order -> dataset disease order -> number
    slug_order = {}
    for d in DISEASES:
        slug_order.setdefault(d["plant"], []).append(d["slug_disease"])
    for p in plant_order:
        slug_order.setdefault(p, []).append("healthy")

    def sort_key(r):
        plist = slug_order[r["plant"]]
        di = plist.index(r["slug"]) if r["slug"] in plist else 999
        return (plant_order.index(r["plant"]), di, int(r["num"]))

    raw.sort(key=sort_key)

    IMG_OUT.mkdir(parents=True, exist_ok=True)
    for i, r in enumerate(raw, start=1):
        plant_slug = PLANT_SLUGS[r["plant"]]
        new_name = f"{plant_slug}_{r['slug']}_{r['num']}.{r['ext']}"
        rel_path = f"images/{plant_slug}/{new_name}"
        dest = OUT / rel_path
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(r["original"], dest)
        review = REVIEW_NOTES.get((r["folder"], r["orig_name"].rsplit(".", 1)[0].rsplit("_", 1)[0], r["num"]))
        # lookup review by original base name instead (safer for spaced names)
        orig_base = re.sub(r"_\d{2}$", "", r["orig_name"].rsplit(".", 1)[0])
        review = None
        for (fld, key, num), note in REVIEW_NOTES.items():
            if r["folder"] == fld and r["num"] == num and orig_base == f"{fld}_{key}":
                review = note
        image_records.append({
            "id": f"IMG{i:03d}", "path": rel_path, "original": r["orig_name"],
            "folder": r["folder"], "plant": r["plant"], "slug": r["slug"],
            "num": r["num"], "status": "needs_review" if review else "available",
            "review": review or "",
        })
else:
    print("WARNING: demo photos folder not found - no images copied")

# group images per plant+slug
from collections import defaultdict
images_by = defaultdict(list)
for rec in image_records:
    images_by[(rec["plant"], rec["slug"])].append(rec)

# ---------------------------------------------------------------- build JSON
def img_entry(rec, label):
    e = {"id": rec["id"], "path": rec["path"], "label": label, "status": rec["status"]}
    if rec["review"]:
        e["review_note"] = rec["review"]
    return e

plants_json = []
for p in plant_order:
    meta = PLANTS[p]
    plant_node = {
        "plant_id": plant_ids[p],
        "name": {"en": p, "ur": meta["ur"]},
        "scientific_name": meta["sci"],
        "category": {"en": meta["category"], "ur": meta["category_ur"]},
        "aliases": meta["aliases"],
        "diseases": [],
    }
    for d in [x for x in DISEASES if x["plant"] == p]:
        recs = images_by.get((p, d["slug_disease"]), [])
        disease_node = {
            "disease_id": d["disease_id"],
            "name": {"en": d["disease"], "ur": d["disease_ur"]},
            "aliases": d["aliases"],
            "type": d["type"],
            "causal_agent": d["causal_agent"],
            "description": {"en": d["description_en"], "ur": d["description_ur"]},
            "symptoms": {"en": d["symptoms_en"], "ur": d["symptoms_ur"]},
            "causes": {"en": d["causes_en"], "ur": d["causes_ur"]},
            "treatment": {"en": d["treatment_en"], "ur": d["treatment_ur"]},
            "prevention": {"en": d["prevention_en"], "ur": d["prevention_ur"]},
            "images": [img_entry(r, d["image_label_en"]) for r in recs],
            "source": {"name": d["source"], "url": d["source_url"]},
            "confidence": d["confidence"],
            "confidence_notes": d["notes"],
        }
        plant_node["diseases"].append(disease_node)
    # healthy entry
    recs = images_by.get((p, "healthy"), [])
    ur_desc, ur_tip = HEALTHY[p]
    healthy_node = {
        "disease_id": HEALTHY_IDS[p],
        "name": {"en": "Healthy (No Disease Detected)", "ur": "صحت مند (کوئی بیماری نہیں)"},
        "aliases": ["healthy"],
        "type": "healthy",
        "causal_agent": None,
        "description": {
            "en": f"The {p.lower()} plant in the image appears healthy - no disease symptoms are visible.",
            "ur": ur_desc,
        },
        "symptoms": {
            "en": ["No disease symptoms visible on the plant in this image."],
            "ur": ["اس تصویر میں پودے پر کسی بیماری کی علامت نظر نہیں آ رہی۔"],
        },
        "causes": {
            "en": "Not applicable - no disease detected.",
            "ur": "لاگو نہیں ہوتا — کوئی بیماری نہیں پائی گئی۔",
        },
        "treatment": {
            "en": ["No treatment needed. Continue good crop care practices."],
            "ur": ["کسی علاج کی ضرورت نہیں۔ فصل کا معمول کے مطابق خیال جاری رکھیں۔"],
        },
        "prevention": {
            "en": ["Inspect the crop weekly so any disease is caught at its earliest stage."],
            "ur": [ur_tip],
        },
        "images": [img_entry(r, f"Healthy {p.lower()} reference photo") for r in recs],
        "source": {"name": "User's reference photo collection", "url": ""},
        "confidence": "high",
        "confidence_notes": "Added so the website can handle a 'healthy' prediction from the classifier; general care guidance only.",
    }
    plant_node["diseases"].append(healthy_node)
    plants_json.append(plant_node)

json_doc = {
    "meta": {
        "dataset_name": "Pakistan Crop Disease Knowledge Base",
        "version": "1.0",
        "generated_on": "2026-08-30",
        "languages": ["en", "ur"],
        "encoding": "UTF-8",
        "plant_count": len(plants_json),
        "disease_count": len(DISEASES),
        "healthy_entries": len(plant_order),
        "image_count": len(image_records),
        "original_source_file": "crop_knowledge_data.xlsx - sheet 'Crop Knowledge' (49 rows, 12 crops)",
        "disclaimer": "This information is for educational purposes and a hackathon prototype. Predictions made by the website are not 100% certain. Always confirm the diagnosis with local agricultural extension staff before taking action. Where a pesticide or fungicide is mentioned, always follow the product label and local agricultural guidance - never exceed the recommended dose.",
        "disclaimer_ur": "یہ معلومات تعلیمی مقاصد اور ہیکاتھون پروٹوٹائپ کے لیے ہیں۔ ویب سائٹ کی جانب سے دی گئی تشخیص سو فیصد درست نہیں ہوتی۔ کسی بھی اقدام سے پہلے اپنے مقامی زرعی ماہر (ایکسٹینشن ورکر) سے تصدیق ضرور کریں۔ جہاں کسی کیڑے مار یا فنگس مار دوا کا ذکر ہے، وہاں ہمیشہ پروڈکٹ کے لیبل اور مقامی زرعی ہدایات پر عمل کریں — تجویز کردہ خوراک سے زیادہ دوا کبھی نہ دیں۔",
    },
    "plants": plants_json,
}

OUT.mkdir(parents=True, exist_ok=True)
with open(OUT / "crop_knowledge.json", "w", encoding="utf-8") as f:
    json.dump(json_doc, f, ensure_ascii=False, indent=2)
print(f"JSON written: {len(plants_json)} plants, {len(DISEASES)} diseases, {len(image_records)} images")

# ---------------------------------------------------------------- image mapping CSV
with open(OUT / "image_mapping.csv", "w", encoding="utf-8-sig", newline="") as f:
    w = csv.writer(f)
    w.writerow(["image_id", "image_path", "original_filename", "source_folder",
                "plant_id", "plant", "disease_id", "disease", "image_label", "status", "review_note"])
    disease_label = {}
    for d in DISEASES:
        disease_label[(d["plant"], d["slug_disease"])] = (d["disease_id"], d["disease"], d["image_label_en"])
    for rec in image_records:
        if rec["slug"] == "healthy":
            did, dname, label = HEALTHY_IDS[rec["plant"]], "Healthy (No Disease Detected)", f"Healthy {rec['plant'].lower()} reference photo"
        else:
            did, dname, label = disease_label[(rec["plant"], rec["slug"])]
        w.writerow([rec["id"], rec["path"], rec["original"], rec["folder"],
                    plant_ids[rec["plant"]], rec["plant"], did, dname, label, rec["status"], rec["review"]])
print("CSV written")

# ---------------------------------------------------------------- Excel
def bullets(items):
    return "\n".join("• " + i for i in items)

wb = openpyxl.Workbook()

# Sheet 1: Dataset
ws = wb.active
ws.title = "Dataset"
headers = [
    "plant_id", "plant_name", "plant_name_urdu", "scientific_name", "category",
    "disease_id", "disease_name", "disease_name_urdu", "disease_type", "causal_agent",
    "aliases", "description", "description_urdu",
    "symptoms", "symptoms_urdu", "causes", "causes_urdu",
    "treatment", "treatment_urdu", "prevention", "prevention_urdu",
    "image_paths", "image_label", "image_count",
    "source_name", "source_url", "confidence", "confidence_notes",
]
ws.append(headers)
hdr_fill = PatternFill("solid", fgColor="2E7D32")
hdr_font = Font(bold=True, color="FFFFFF")
for c in ws[1]:
    c.fill = hdr_fill
    c.font = hdr_font
    c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

for p in plant_order:
    meta = PLANTS[p]
    for d in [x for x in DISEASES if x["plant"] == p]:
        recs = images_by.get((p, d["slug_disease"]), [])
        paths = "\n".join(r["path"] for r in recs)
        ws.append([
            plant_ids[p], p, meta["ur"], meta["sci"], meta["category"],
            d["disease_id"], d["disease"], d["disease_ur"], d["type"], d["causal_agent"],
            ", ".join(d["aliases"]), d["description_en"], d["description_ur"],
            bullets(d["symptoms_en"]), bullets(d["symptoms_ur"]),
            d["causes_en"], d["causes_ur"],
            bullets(d["treatment_en"]), bullets(d["treatment_ur"]),
            bullets(d["prevention_en"]), bullets(d["prevention_ur"]),
            paths, d["image_label_en"], len(recs),
            d["source"], d["source_url"], d["confidence"], d["notes"],
        ])
    # healthy row
    recs = images_by.get((p, "healthy"), [])
    ur_desc, ur_tip = HEALTHY[p]
    ws.append([
        plant_ids[p], p, meta["ur"], meta["sci"], meta["category"],
        HEALTHY_IDS[p], "Healthy (No Disease Detected)", "صحت مند (کوئی بیماری نہیں)", "healthy", "",
        "healthy",
        f"The {p.lower()} plant appears healthy - no disease symptoms are visible.", ur_desc,
        "• No disease symptoms visible on the plant.", "• اس تصویر میں پودے پر کسی بیماری کی علامت نظر نہیں آ رہی۔",
        "Not applicable - no disease detected.", "لاگو نہیں ہوتا — کوئی بیماری نہیں پائی گئی۔",
        "• No treatment needed. Continue good crop care practices.",
        "• کسی علاج کی ضرورت نہیں۔ فصل کا معمول کے مطابق خیال جاری رکھیں۔",
        "• Inspect the crop weekly so any disease is caught at its earliest stage.", "• " + ur_tip,
        "\n".join(r["path"] for r in recs), f"Healthy {p.lower()} reference photo", len(recs),
        "User's reference photo collection", "", "high",
        "Added so the website can handle a 'healthy' prediction from the classifier.",
    ])

widths = [9, 12, 16, 22, 11, 10, 26, 26, 11, 28, 18, 45, 45, 45, 45, 45, 45, 45, 45, 45, 45, 34, 30, 7, 30, 40, 14, 55]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for row in ws.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True)
ws.freeze_panes = "G2"

# Sheet 2: Image Mapping
ws2 = wb.create_sheet("Image Mapping")
ws2.append(["image_id", "image_path", "original_filename", "source_folder",
            "plant_id", "plant", "disease_id", "disease", "image_label", "status", "review_note"])
for c in ws2[1]:
    c.fill = hdr_fill
    c.font = hdr_font
    c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
disease_label = {}
for d in DISEASES:
    disease_label[(d["plant"], d["slug_disease"])] = (d["disease_id"], d["disease"], d["image_label_en"])
for rec in image_records:
    if rec["slug"] == "healthy":
        did, dname, label = HEALTHY_IDS[rec["plant"]], "Healthy (No Disease Detected)", f"Healthy {rec['plant'].lower()} reference photo"
    else:
        did, dname, label = disease_label[(rec["plant"], rec["slug"])]
    ws2.append([rec["id"], rec["path"], rec["original"], rec["folder"],
                plant_ids[rec["plant"]], rec["plant"], did, dname, label, rec["status"], rec["review"]])
for i, w in enumerate([10, 44, 38, 12, 9, 12, 10, 28, 42, 13, 60], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
for row in ws2.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True)
ws2.freeze_panes = "A2"

# Sheet 3: Quality Log
ws3 = wb.create_sheet("Quality Log")
ws3.append(["#", "scope", "item", "issue", "action taken / recommendation"])
for c in ws3[1]:
    c.fill = PatternFill("solid", fgColor="B71C1C")
    c.font = hdr_font
    c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

quality_log = [
    ("global", "Column structure", "Original sheet had only 6 columns (Crop, Disease, Symptoms, Basic Advice, Source, URL); no causes, prevention, descriptions, or Urdu.", "Expanded to 28 columns incl. Urdu columns for every user-facing field, IDs, image mapping, and confidence notes."),
    ("global", "Mojibake (corrupted characters)", "7 rows contained a corrupted character (likely an en-dash) e.g. 'No cure once infected \ufffd remove...'.", "Replaced with correct punctuation in the rewritten text."),
    ("global", "Source name typos", "'planttix' -> Plantix; 'UthaStateUniversity' -> Utah State University; 'UtahState-University' -> Utah State University; 'SASKATCHEWAN Pulse Growers' -> Saskatchewan Pulse Growers; 'Wordpress' -> La Riceman (WordPress blog).", "Standardised source names."),
    ("global", "Spelling/capitalisation", "'Alternaria leaf Spot' -> 'Alternaria Leaf Spot'; 'Styler End Rot' -> 'Stylar End Rot'; 'Die Back' -> renamed to 'Sooty Mold' (see below).", "Standardised disease names."),
    ("global", "Treatment vs prevention mixed", "The single 'Basic Advice' column mixed immediate treatment and long-term prevention.", "Split into separate 'treatment' and 'prevention' fields, each in English and Urdu."),
    ("global", "Missing causes", "No causal agents were given in the original sheet.", "Added causal organisms from standard plant pathology references; uncertain ones marked for verification."),
    ("global", "No Urdu", "The original sheet had no Urdu.", "All user-facing fields translated into natural Urdu for Pakistani users; technical terms kept alongside Urdu where clearer."),
    ("disease", "Mango 'Die Back' row", "Symptoms and advice in the original row describe a black soot-like coating from insect honeydew - this is SOOTY MOLD, not dieback (which is gradual twig death). The user's own photos are named 'Mango_Sooty_Mold'.", "Renamed to 'Sooty Mold', kept 'die back' as alias, flagged for manual confirmation."),
    ("disease", "Onion Downy Mildew", "No reference photo in the collected photo set.", "Disease kept (from original sheet); image list empty - add a photo."),
    ("disease", "Onion Powdery Mildew", "Two photos ('Onion_Powdery_Mildew_01/02') exist but the disease was not in the original Excel sheet.", "Added as a new entry (DS033) marked 'needs_verification'; content written from standard pathology knowledge."),
    ("disease", "Chili Anthracnose image", "The single photo shows damaged leaves (not fruit rot) and has a Shutterstock watermark.", "Image linked but marked 'needs_review' - replace before production. Original sheet's URL was also a Shutterstock link rather than the cited Frontiers article."),
    ("disease", "Onion Basal Rot image 01", "Sampled photo shows garlic, not onion.", "Marked 'needs_review' - replace or relabel."),
    ("disease", "Guava Fruit Scab image", "Sampled photo appears to be an anthracnose diagnostic collage (Colletotrichum spores), not corky scab.", "Marked 'needs_review' - verify or replace. Causal agent of guava fruit scab also needs verification."),
    ("disease", "Wheat Mildew image 01", "Sampled photo shows a wetter/standing-water setting atypical of wheat.", "Marked 'needs_review' - verify the crop shown."),
    ("disease", "Citrus Greening image 01", "Sampled photo shows normal green fruit with no clear HLB symptoms.", "Marked 'needs_review' - verify or replace."),
    ("disease", "Cotton source URLs", "3 cotton rows had URLs ending with a placeholder note 'TBD - no confirmed dataset found...'.", "Trimmed to the valid Cotton Incorporated URL; the note is preserved in confidence_notes."),
    ("disease", "Onion Downy Mildew URL", "URL is a search-results page rather than an article.", "Kept; replace with a direct extension page when available."),
    ("disease", "Guava stylar end rot / fruit scab causal agents", "Causal organisms not confirmed from available sources.", "Described as fungal with 'needs verification' confidence; confirm with a plant pathology source."),
    ("image", "File naming", "Original photo filenames had spaces ('Cotton_Ascochyta Blight_01.jpg'), parentheses ('Chilli_Anthracnose (Fruit Rot)_01.webp'), mixed case ('Gray_leaf_spot', 'leaf_Smut'), a typo ('Styler_End_Root'), and mixed extensions (.jfif, .JPG, .webp, .png).", "All 116 photos copied with normalised names: images/{plant}/{plant}_{disease}_{nn}.{ext}; .jfif/.JPG renamed to .jpg; webp/png kept (browser-compatible)."),
    ("image", "Verification coverage", "Only ~10% of photos (12 of 116) were visually spot-checked.", "Most matched labels; 5 problems found and marked 'needs_review'. Recommend reviewing all images before production."),
    ("image", "Healthy photos", "Each crop folder contains one 'Healthy' photo, but the Excel had no healthy rows.", "Added a bilingual 'Healthy' entry per crop (DS051-DS062) so the site can handle healthy predictions."),
]
for i, (scope, item, issue, action) in enumerate(quality_log, start=1):
    ws3.append([i, scope, item, issue, action])
for i, w in enumerate([5, 10, 26, 62, 62], start=1):
    ws3.column_dimensions[get_column_letter(i)].width = w
for row in ws3.iter_rows(min_row=2):
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True)
ws3.freeze_panes = "A2"

wb.save(OUT / "crop_knowledge_enhanced.xlsx")
print("Excel written")

# ---------------------------------------------------------------- summary
disease_slugs = {(d["plant"], d["slug_disease"]) for d in DISEASES}
mapped_pairs = {(r["plant"], r["slug"]) for r in image_records}
no_images = [f"{p} - {s}" for (p, s) in disease_slugs - mapped_pairs]
print("\nDiseases without images:", no_images or "none")
print("Unmapped photo files:", unmapped_files or "none")
print("Images needing review:", sum(1 for r in image_records if r["status"] == "needs_review"))
print("Total images copied:", len(image_records))

# -*- coding: utf-8 -*-
"""Validate all generated outputs. Usage: python validate_outputs.py"""
import csv
import json
import re
from pathlib import Path

import openpyxl

BASE = Path(__file__).parent
OUT = BASE / "plant_disease_dataset"
errors, warnings = [], []

# ---------------------------------------------------------------- JSON
raw = (OUT / "crop_knowledge.json").read_bytes()
if raw.startswith(b"\xef\xbb\xbf"):
    errors.append("JSON has a BOM - should be plain UTF-8")
doc = json.loads(raw.decode("utf-8"))

meta = doc["meta"]
plants = doc["plants"]
if meta["plant_count"] != len(plants):
    errors.append(f"meta.plant_count {meta['plant_count']} != {len(plants)}")
total_diseases = sum(len(p["diseases"]) for p in plants)
if meta["disease_count"] != total_diseases - 12:
    errors.append(f"meta.disease_count mismatch: {meta['disease_count']} vs {total_diseases - 12} actual")
if meta["healthy_entries"] != 12:
    errors.append("healthy_entries != 12")
if not meta.get("disclaimer") or not meta.get("disclaimer_ur"):
    errors.append("meta disclaimer missing in EN or UR")

plant_ids, disease_ids, image_ids, image_paths = [], [], [], []
ur_pat = re.compile(r"[\u0600-\u06FF]")
mojibake_pat = re.compile(r"[\ufffd\u00c2\u00e2]")

REQUIRED_BILINGUAL = ["name", "description", "symptoms", "causes", "treatment", "prevention"]
LIST_FIELDS = ["symptoms", "treatment", "prevention"]

for p in plants:
    pid = p.get("plant_id")
    if not pid or pid in plant_ids:
        errors.append(f"bad/duplicate plant_id {pid}")
    plant_ids.append(pid)
    for f in ["name", "category"]:
        if not p[f]["en"] or not p[f]["ur"]:
            errors.append(f"{pid}: {f} empty in en/ur")
    if not ur_pat.search(p["name"]["ur"]):
        errors.append(f"{pid}: plant name ur not Urdu script")
    for d in p["diseases"]:
        did = d.get("disease_id")
        if not did or did in disease_ids:
            errors.append(f"bad/duplicate disease_id {did} ({p['name']['en']})")
        disease_ids.append(did)
        for f in REQUIRED_BILINGUAL:
            node = d.get(f)
            if not isinstance(node, dict) or "en" not in node or "ur" not in node:
                errors.append(f"{did}: {f} not a bilingual object")
                continue
            for lang in ["en", "ur"]:
                v = node[lang]
                if isinstance(v, list):
                    if not v or any(not str(x).strip() for x in v):
                        errors.append(f"{did}: {f}.{lang} empty item")
                elif not str(v).strip():
                    errors.append(f"{did}: {f}.{lang} empty string")
                if mojibake_pat.search(str(v)):
                    errors.append(f"{did}: {f}.{lang} contains mojibake")
        # parallel arrays
        for f in LIST_FIELDS:
            if len(d[f]["en"]) != len(d[f]["ur"]):
                errors.append(f"{did}: {f} en/ur length mismatch")
        # Urdu script present in name ur
        if not ur_pat.search(d["name"]["ur"]):
            errors.append(f"{did}: disease name ur not Urdu script: {d['name']['ur']}")
        # Urdu fields actually contain Urdu
        if not ur_pat.search(d["description"]["ur"]):
            errors.append(f"{did}: description.ur has no Urdu script")
        if d["type"] not in ("fungal", "bacterial", "viral", "pest", "physiological", "algal", "healthy"):
            errors.append(f"{did}: unexpected type {d['type']}")
        if d["confidence"] not in ("high", "medium", "needs_verification"):
            errors.append(f"{did}: unexpected confidence {d['confidence']}")
        if not d.get("confidence_notes"):
            errors.append(f"{did}: confidence_notes empty")
        if d["type"] != "healthy" and not d.get("causal_agent"):
            errors.append(f"{did}: causal_agent empty")
        for img in d["images"]:
            iid = img.get("id")
            if not iid or iid in image_ids:
                errors.append(f"bad/duplicate image id {iid}")
            image_ids.append(iid)
            image_paths.append(img["path"])
            full = OUT / img["path"]
            if not full.is_file():
                errors.append(f"{iid}: file missing on disk: {img['path']}")
            if img.get("status") not in ("available", "needs_review"):
                errors.append(f"{iid}: bad status {img.get('status')}")
            if img.get("status") == "needs_review" and not img.get("review_note"):
                errors.append(f"{iid}: needs_review without review_note")

if meta["image_count"] != len(image_ids):
    errors.append(f"meta.image_count {meta['image_count']} != {len(image_ids)}")

# unique image file paths
if len(set(image_paths)) != len(image_paths):
    dupes = {x for x in image_paths if image_paths.count(x) > 1}
    errors.append(f"duplicate image paths: {dupes}")

# duplicates of (plant, disease)
pairs = [(p["name"]["en"], d["name"]["en"]) for p in plants for d in p["diseases"]]
if len(set(pairs)) != len(pairs):
    errors.append("duplicate (plant, disease) pairs exist")

# ---------------------------------------------------------------- CSV
with open(OUT / "image_mapping.csv", encoding="utf-8-sig") as f:
    rows = list(csv.DictReader(f))
if len(rows) != 116:
    errors.append(f"CSV has {len(rows)} rows, expected 116")
csv_ids = [r["image_id"] for r in rows]
if sorted(csv_ids) != sorted(image_ids):
    errors.append("CSV image ids differ from JSON image ids")
for r in rows:
    if mojibake_pat.search(r["review_note"]):
        errors.append(f"CSV {r['image_id']}: mojibake in review_note")
    if not (OUT / r["image_path"]).is_file():
        errors.append(f"CSV {r['image_id']}: file missing {r['image_path']}")
statuses = {r["status"] for r in rows}
if statuses - {"available", "needs_review"}:
    errors.append(f"CSV unexpected statuses {statuses}")

# ---------------------------------------------------------------- Excel
wb = openpyxl.load_workbook(OUT / "crop_knowledge_enhanced.xlsx")
if wb.sheetnames != ["Dataset", "Image Mapping", "Quality Log"]:
    errors.append(f"Excel sheets: {wb.sheetnames}")

ws = wb["Dataset"]
if ws.max_row != 63:  # header + 50 diseases + 12 healthy
    errors.append(f"Dataset sheet has {ws.max_row} rows, expected 63")
if ws.max_column != 28:
    errors.append(f"Dataset sheet has {ws.max_column} cols, expected 28")
hdr = [c.value for c in ws[1]]
expected_hdr = ["plant_id", "plant_name", "plant_name_urdu", "scientific_name", "category",
                "disease_id", "disease_name", "disease_name_urdu", "disease_type", "causal_agent",
                "aliases", "description", "description_urdu", "symptoms", "symptoms_urdu",
                "causes", "causes_urdu", "treatment", "treatment_urdu", "prevention",
                "prevention_urdu", "image_paths", "image_label", "image_count", "source_name",
                "source_url", "confidence", "confidence_notes"]
if hdr != expected_hdr:
    errors.append("Dataset header mismatch")
xl_ids = [ws.cell(row=r, column=6).value for r in range(2, ws.max_row + 1)]
if xl_ids != disease_ids:
    errors.append("Excel disease_id column differs from JSON order")
for r in range(2, ws.max_row + 1):
    for c in range(1, 29):
        v = ws.cell(row=r, column=c).value
        if v and mojibake_pat.search(str(v)):
            errors.append(f"Excel Dataset r{r}c{c}: mojibake")
    # every non-healthy row must have non-empty core fields
    if ws.cell(row=r, column=9).value != "healthy":
        for c in [2, 3, 7, 8, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 27]:
            if not str(ws.cell(row=r, column=c).value or "").strip():
                errors.append(f"Excel Dataset r{r} col {expected_hdr[c-1]} empty")

ws2 = wb["Image Mapping"]
if ws2.max_row != 117:
    errors.append(f"Image Mapping sheet has {ws2.max_row} rows, expected 117")
ws3 = wb["Quality Log"]
if ws3.max_row < 22:
    errors.append(f"Quality Log has only {ws3.max_row} rows")

# ---------------------------------------------------------------- image files on disk
img_files = list((OUT / "images").rglob("*"))
img_files = [f for f in img_files if f.is_file()]
if len(img_files) != 116:
    errors.append(f"disk images {len(img_files)} != 116")
bad_names = [f.name for f in img_files if not re.match(r"^[a-z0-9_]+_(healthy|[a-z0-9_]+)_\d{2}\.(jpg|png|webp)$", f.name)]
if bad_names:
    errors.append(f"non-normalised image names: {bad_names[:5]}")
bad_ext = [f.name for f in img_files if f.suffix.lower() in (".jfif", ".jpeg")]
if bad_ext:
    errors.append(f"unnormalised extensions: {bad_ext}")

# source folder untouched
src = Path(r"C:\Users\DELL\Downloads\demo photos")
src_count = sum(1 for f in src.rglob("*") if f.is_file())
if src_count != 116:
    warnings.append(f"source folder now has {src_count} files (expected 116)")

print(f"Plants: {len(plants)} | disease entries: {len(disease_ids)} (50 diseases + 12 healthy) | images: {len(image_ids)}")
print(f"JSON size: {raw.count(chr(10).encode())} lines, {len(raw)/1024:.0f} KB")
print(f"Confidence distribution:", {c: sum(1 for p in plants for d in p['diseases'] if d['confidence'] == c) for c in ['high', 'medium', 'needs_verification']})
if warnings:
    print("WARNINGS:", warnings)
if errors:
    print(f"\nFAILED - {len(errors)} error(s):")
    for e in errors[:40]:
        print(" -", e)
    raise SystemExit(1)
print("\nALL CHECKS PASSED")
